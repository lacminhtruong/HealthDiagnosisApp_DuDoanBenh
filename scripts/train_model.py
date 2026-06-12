from __future__ import annotations

import argparse
import json
import random
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import torch
from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATASET_PATH = PROJECT_ROOT / "datasets" / "massive_health_training_data_v2.xlsx"
MODEL_PATH = PROJECT_ROOT / "trained_models" / "health_diagnosis_model.pt"
META_PATH = PROJECT_ROOT / "trained_models" / "health_diagnosis_model_meta.json"
SYMPTOM_PREFIX = "symptom__"
COLUMN_ALIASES = {
    "disease": ("disease", "Bệnh"),
    "age": ("age", "Tuổi"),
    "gender": ("gender", "Giới tính"),
    "height_cm": ("height_cm", "Chiều cao (cm)"),
    "weight_kg": ("weight_kg", "Cân nặng (kg)"),
    "bmi": ("bmi", "BMI"),
}

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    return " ".join(unicodedata.normalize("NFC", text).strip().split())

@dataclass
class DatasetMatrix:
    features: torch.Tensor
    labels: torch.Tensor
    feature_columns: list[str]
    disease_to_label: dict[str, int]
    dataset_path: Path
    normalization_params: dict[str, float]

@dataclass
class DatasetSplit:
    train_indices: list[int]
    test_indices: list[int]
    valid_indices: list[int]


def find_column_index(headers: list[object], field_name: str, required: bool = True) -> int | None:
    normalized_headers = {
        normalize_text(header).casefold(): index
        for index, header in enumerate(headers)
        if header is not None
    }
    for alias in COLUMN_ALIASES[field_name]:
        index = normalized_headers.get(normalize_text(alias).casefold())
        if index is not None:
            return index

    if required:
        aliases = ", ".join(COLUMN_ALIASES[field_name])
        raise ValueError(f"Dataset is missing required column '{field_name}' (accepted names: {aliases}).")
    return None


def parse_number(value: object, field_name: str, row_number: int) -> float | None:
    if value is None or normalize_text(value) == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Invalid numeric value for '{field_name}' at Excel row {row_number}: {value!r}"
        ) from exc


def encode_gender(value: object, row_number: int) -> float:
    normalized = normalize_text(value).casefold()
    if normalized in {"nữ", "nu", "female", "f", "1"}:
        return 1.0
    if normalized in {"nam", "male", "m", "0"}:
        return 0.0
    raise ValueError(f"Invalid gender at Excel row {row_number}: {value!r}")


def min_max_scale(value: float, minimum: float, maximum: float) -> float:
    if maximum == minimum:
        return 0.0
    return (value - minimum) / (maximum - minimum)


def read_excel_dataset(path: Path = DATASET_PATH) -> DatasetMatrix:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["training_data"] if "training_data" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Excel dataset does not contain training rows.")

    headers = list(rows[0])
    column_indices = {
        field_name: find_column_index(headers, field_name, required=field_name not in {"age", "gender"})
        for field_name in COLUMN_ALIASES
    }
    if (column_indices["age"] is None) != (column_indices["gender"] is None):
        raise ValueError("Dataset must contain both age and gender columns, or neither.")

    symptom_indices = [
        index
        for index, header in enumerate(headers)
        if normalize_text(header).casefold().startswith(SYMPTOM_PREFIX)
    ]
    if not symptom_indices:
        raise ValueError(f"Dataset does not contain any '{SYMPTOM_PREFIX}' feature columns.")

    symptom_feature_names = [
        f"{SYMPTOM_PREFIX}{normalize_text(headers[index])[len(SYMPTOM_PREFIX):]}"
        for index in symptom_indices
    ]
    if len(symptom_feature_names) != len(set(symptom_feature_names)):
        raise ValueError("Dataset contains duplicate symptom feature columns.")

    numeric_fields = ["height_cm", "weight_kg", "bmi"]
    if column_indices["age"] is not None:
        numeric_fields.insert(0, "age")

    prepared_rows: list[dict[str, object]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        disease_index = column_indices["disease"]
        assert disease_index is not None
        disease = normalize_text(row[disease_index])
        if not disease:
            continue

        numeric_values = {}
        for field_name in numeric_fields:
            field_index = column_indices[field_name]
            assert field_index is not None
            numeric_values[field_name] = parse_number(row[field_index], field_name, row_number)

        gender_index = column_indices["gender"]
        prepared_rows.append(
            {
                "disease": disease,
                "numeric_values": numeric_values,
                "gender_encoded": encode_gender(row[gender_index], row_number) if gender_index is not None else None,
                "symptoms": [
                    parse_number(row[index], normalize_text(headers[index]), row_number) or 0.0
                    for index in symptom_indices
                ],
            }
        )

    if not prepared_rows:
        raise ValueError("Excel dataset does not contain usable training rows.")

    numeric_means: dict[str, float] = {}
    normalization_params: dict[str, float] = {}
    for field_name in numeric_fields:
        values = [
            row["numeric_values"][field_name]
            for row in prepared_rows
            if row["numeric_values"][field_name] is not None
        ]
        if not values:
            raise ValueError(f"Column '{field_name}' does not contain any numeric values.")

        numeric_means[field_name] = sum(values) / len(values)
        filled_values = [
            row["numeric_values"][field_name]
            if row["numeric_values"][field_name] is not None
            else numeric_means[field_name]
            for row in prepared_rows
        ]
        normalization_params[f"{field_name}_min"] = float(min(filled_values))
        normalization_params[f"{field_name}_max"] = float(max(filled_values))
        normalization_params[f"{field_name}_default"] = float(numeric_means[field_name])

    base_feature_columns = [f"{field_name}_scaled" for field_name in numeric_fields]
    if column_indices["gender"] is not None:
        base_feature_columns.insert(1, "gender_encoded")
    feature_columns = [*base_feature_columns, *symptom_feature_names]

    diseases: list[str] = []
    feature_rows: list[list[float]] = []
    label_rows: list[int] = []
    for row in prepared_rows:
        disease = row["disease"]
        if disease not in diseases:
            diseases.append(disease)

        numeric_values = row["numeric_values"]
        features = []
        for field_name in numeric_fields:
            value = numeric_values[field_name]
            if value is None:
                value = numeric_means[field_name]
            features.append(
                min_max_scale(
                    value,
                    normalization_params[f"{field_name}_min"],
                    normalization_params[f"{field_name}_max"],
                )
            )
            if field_name == "age" and row["gender_encoded"] is not None:
                features.append(row["gender_encoded"])
        features.extend(row["symptoms"])
        feature_rows.append(features)

    disease_to_label = {disease: index for index, disease in enumerate(diseases)}
    label_rows = [disease_to_label[row["disease"]] for row in prepared_rows]

    return DatasetMatrix(
        features=torch.tensor(feature_rows, dtype=torch.float32),
        labels=torch.tensor(label_rows, dtype=torch.long),
        feature_columns=feature_columns,
        disease_to_label=disease_to_label,
        dataset_path=path,
        normalization_params=normalization_params,
    )

def stratified_split(labels: torch.Tensor, seed: int = 12) -> DatasetSplit:
    rng = random.Random(seed)
    label_to_indices: dict[int, list[int]] = defaultdict(list)

    for index, label in enumerate(labels.tolist()):
        label_to_indices[int(label)].append(index)

    train_indices: list[int] = []
    test_indices: list[int] = []
    valid_indices: list[int] = []

    for indices in label_to_indices.values():
        rng.shuffle(indices)
        count = len(indices)

        if count >= 3:
            train_count = max(1, int(count * 0.7))
            test_count = max(1, int(count * 0.1))
            valid_count = count - train_count - test_count
            if valid_count < 1:
                train_count -= 1
                valid_count = 1
        elif count == 2:
            train_count = 1
            test_count = 1
            valid_count = 0
        else:
            train_count = 1
            test_count = 0
            valid_count = 0

        train_indices.extend(indices[:train_count])
        test_indices.extend(indices[train_count:train_count + test_count])
        valid_indices.extend(indices[train_count + test_count:train_count + test_count + valid_count])

    rng.shuffle(train_indices)
    rng.shuffle(test_indices)
    rng.shuffle(valid_indices)

    return DatasetSplit(
        train_indices=train_indices,
        test_indices=test_indices,
        valid_indices=valid_indices,
    )

def predict_knn(
    query_features: torch.Tensor,
    train_features: torch.Tensor,
    train_labels: torch.Tensor,
    k: int,
    batch_size: int = 256,
) -> torch.Tensor:
    predictions: list[int] = []
    effective_k = min(k, train_features.shape[0])

    for start in range(0, query_features.shape[0], batch_size):
        batch = query_features[start:start + batch_size]
        distances = torch.cdist(batch, train_features)
        neighbor_distances, neighbor_indices = torch.topk(
            distances,
            k=effective_k,
            largest=False,
        )
        neighbor_labels = train_labels[neighbor_indices]

        for sample_labels, sample_distances in zip(neighbor_labels, neighbor_distances):
            label_votes = Counter(sample_labels.tolist())
            best_vote_count = max(label_votes.values())
            candidates = [
                label
                for label, vote_count in label_votes.items()
                if vote_count == best_vote_count
            ]
            if len(candidates) == 1:
                predictions.append(candidates[0])
                continue

            best_label = min(
                candidates,
                key=lambda label: average_distance_for_label(
                    sample_labels,
                    sample_distances,
                    label,
                ),
            )
            predictions.append(best_label)

    return torch.tensor(predictions, dtype=torch.long)

def average_distance_for_label(labels: torch.Tensor, distances: torch.Tensor, label: int) -> float:
    matching_distances = [
        distance
        for current_label, distance in zip(labels.tolist(), distances.tolist())
        if current_label == label
    ]
    return sum(matching_distances) / len(matching_distances)

def accuracy_score(predictions: torch.Tensor, labels: torch.Tensor) -> float:
    if labels.numel() == 0:
        return 0.0
    return (predictions == labels).float().mean().item()

def evaluate_split(
    features: torch.Tensor,
    labels: torch.Tensor,
    train_features: torch.Tensor,
    train_labels: torch.Tensor,
    k: int,
) -> float:
    if features.shape[0] == 0:
        return 0.0
    predictions = predict_knn(features, train_features, train_labels, k)
    return round(accuracy_score(predictions, labels), 4)

def build_artifact(matrix: DatasetMatrix, split: DatasetSplit, k: int, seed: int) -> dict:
    train_features = matrix.features[split.train_indices]
    train_labels = matrix.labels[split.train_indices]
    test_features = matrix.features[split.test_indices]
    test_labels = matrix.labels[split.test_indices]
    valid_features = matrix.features[split.valid_indices]
    valid_labels = matrix.labels[split.valid_indices]

    label_to_disease = {
        str(label): disease
        for disease, label in matrix.disease_to_label.items()
    }
    metrics = {
        "algorithm": "KNN",
        "k": k,
        "seed": seed,
        "total_rows": matrix.features.shape[0],
        "training_rows": train_features.shape[0],
        "test_rows": test_features.shape[0],
        "valid_rows": valid_features.shape[0],
        "training_accuracy": evaluate_split(train_features, train_labels, train_features, train_labels, k),
        "test_accuracy": evaluate_split(test_features, test_labels, train_features, train_labels, k),
        "valid_accuracy": evaluate_split(valid_features, valid_labels, train_features, train_labels, k),
        "split_ratio": {
            "train": 0.7,
            "test": 0.1,
            "valid": 0.2,
        },
    }

    return {
        "model_type": "knn",
        "algorithm": "KNN",
        "k": k,
        "train_features": train_features,
        "train_labels": train_labels,
        "test_features": test_features,
        "test_labels": test_labels,
        "valid_features": valid_features,
        "valid_labels": valid_labels,
        "feature_columns": matrix.feature_columns,
        "disease_to_label": matrix.disease_to_label,
        "label_to_disease": label_to_disease,
        "metrics": metrics,
        "split_indices": {
            "train": split.train_indices,
            "test": split.test_indices,
            "valid": split.valid_indices,
        },
        "normalization": matrix.normalization_params,
    }

def save_artifact(artifact: dict, matrix: DatasetMatrix, model_path: Path, meta_path: Path) -> Path:
    model_path.parent.mkdir(parents=True, exist_ok=True)
    torch.save(artifact, model_path)

    meta_path.write_text(
        json.dumps(
            {
                "model_path": str(model_path),
                "dataset_path": str(matrix.dataset_path),
                "algorithm": artifact["algorithm"],
                "k": artifact["k"],
                "features": len(matrix.feature_columns),
                "classes": len(matrix.disease_to_label),
                "metrics": artifact["metrics"],
                "disease_to_label": matrix.disease_to_label,
                "normalization": artifact["normalization"],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    return model_path

def resolve_path(path: Path) -> Path:
    return path if path.is_absolute() else PROJECT_ROOT / path

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dataset", type=Path, default=DATASET_PATH)
    parser.add_argument("--model", type=Path, default=MODEL_PATH)
    parser.add_argument("--meta", type=Path, default=META_PATH)
    parser.add_argument("--k", type=int, default=5)
    parser.add_argument("--seed", type=int, default=12)
    args = parser.parse_args()

    dataset_path = resolve_path(args.dataset)
    model_path = resolve_path(args.model)
    meta_path = resolve_path(args.meta)

    if not dataset_path.exists():
        raise FileNotFoundError(f"Dataset not found: {dataset_path}")
    if args.k < 1:
        raise ValueError("K must be greater than or equal to 1.")

    matrix = read_excel_dataset(dataset_path)
    split = stratified_split(matrix.labels, seed=args.seed)
    artifact = build_artifact(matrix, split, args.k, args.seed)
    saved_model_path = save_artifact(artifact, matrix, model_path, meta_path)
    metrics = artifact["metrics"]

    print(f"Dataset: {dataset_path.relative_to(PROJECT_ROOT)}")
    print(f"Algorithm: KNN")
    print(f"K: {args.k}")
    print(f"Rows: {metrics['total_rows']}")
    print(f"Train/Test/Valid: {metrics['training_rows']}/{metrics['test_rows']}/{metrics['valid_rows']}")
    print(f"Classes: {len(matrix.disease_to_label)}")
    print(f"Features: {len(matrix.feature_columns)}")
    print(f"Trained model: {saved_model_path.relative_to(PROJECT_ROOT)}")
    print(f"Train accuracy: {metrics['training_accuracy']}")
    print(f"Test accuracy: {metrics['test_accuracy']}")
    print(f"Valid accuracy: {metrics['valid_accuracy']}")

if __name__ == "__main__":
    main()
