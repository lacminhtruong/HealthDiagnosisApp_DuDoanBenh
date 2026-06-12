from __future__ import annotations

import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from data import SYMPTOMS, TRAINING_DATA  # noqa: E402

OUTPUT_PATH = PROJECT_ROOT / "datasets" / "health_training_data.xlsx"
SYMPTOM_PREFIX = "symptom__"


def normalize_text(value: object) -> str:
    return " ".join(unicodedata.normalize("NFC", str(value or "")).strip().split())


def build_symptom_vocabulary() -> list[str]:
    symptoms = list(SYMPTOMS)
    seen = set(symptoms)

    for sample in TRAINING_DATA:
        for symptom in sample.patient.symptoms:
            normalized_symptom = normalize_text(symptom)
            if normalized_symptom not in seen:
                symptoms.append(normalized_symptom)
                seen.add(normalized_symptom)

    return symptoms


def export_dataset(output_path: Path = OUTPUT_PATH) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    symptoms = build_symptom_vocabulary()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "training_data"

    headers = [
        "disease",
        "height_cm",
        "weight_kg",
        "bmi",
        "symptoms",
        *[f"{SYMPTOM_PREFIX}{symptom}" for symptom in symptoms],
    ]
    sheet.append(headers)

    for sample in TRAINING_DATA:
        patient_symptoms = {
            normalize_text(symptom)
            for symptom in sample.patient.symptoms
        }
        row = [
            normalize_text(sample.disease_name),
            sample.patient.height_cm,
            sample.patient.weight_kg,
            round(sample.patient.bmi, 4),
            ", ".join(normalize_text(symptom) for symptom in sample.patient.symptoms),
            *[1 if symptom in patient_symptoms else 0 for symptom in symptoms],
        ]
        sheet.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="DDEFEA")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sheet.freeze_panes = "A2"
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 12), 34)

    workbook.save(output_path)
    return output_path


if __name__ == "__main__":
    path = export_dataset()
    print(f"Exported dataset: {path.relative_to(PROJECT_ROOT)}")
